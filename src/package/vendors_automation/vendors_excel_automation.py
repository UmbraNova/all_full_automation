import openpyxl
from openpyxl.utils import column_index_from_string


# TODO: doesn't work, missing data... eronate data


def open_excel(file_path, sheet_name=None):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        return workbook, sheet
    except KeyError:
        print(f"Sheet '{sheet_name}' not found in {file_path}.")
        return None, None

def get_cell_value(sheet, row, col_letter):
    # """Get the value from a specific cell."""
    try:
        col = column_index_from_string(col_letter.upper())
        cell = sheet.cell(row=row, column=col)
        return cell.value
    except ValueError:
        print(f"Invalid column letter: {col_letter}")
        return None

def set_cell_value(sheet, row, col_letter, value):
    # """Set a specific cell's value."""
    try:
        col = column_index_from_string(col_letter.upper())
        sheet.cell(row=row, column=col, value=value)
    except ValueError:
        print(f"Invalid column letter: {col_letter}")

def copy_cells(file1_sheet, file2_sheet):
    # Copy multiple cells from file1 to file2
    row_start = int(input("Enter START row file 1: "))
    row_end = int(input("Enter END row file 1: "))
    num_rows = (row_end + 1) - row_start

    for i_row in range(num_rows):
        num_cells_1 = ["G", "H", "I", "J", "K", "L", "M", "Q", "U", "V"]
        num_cells_2 = ["P", "C", "AF", "I", "F", "G", "H", "AD", "AE", "E"]

        # adding indexing for rows
        set_cell_value(file2_sheet, 6 + i_row, "B", i_row + 1)

        for j_column in range(len(num_cells_1)):
            # adding values from file1 to file2
            row_1 = row_start + i_row
            col_1 = num_cells_1[j_column]
            value = get_cell_value(file1_sheet, row_1, col_1)

            row_2 = 6 + i_row
            col_2 = num_cells_2[j_column]

            if value is not None:
                set_cell_value(file2_sheet, row_2, col_2, value)

        add_custom_data(i_row, file2_sheet)


# test//////////////////////////////////////////////////////////////////
def add_custom_data(i_row, file2_sheet):
    # fixed data
    add_to_cell = ["D", "J", "K", "L", "N", "Q" "R", "S", "X", "Y", "AG"]
    values_to_add = ["eng description", "MARFA_19", "TVA19B", "RO", "cod_TODO", "_TODO", "30DAYSSALES", "1", "BUC", "1", "name_TODO"]        
    for j_column in range(len(add_to_cell)):
        value = values_to_add[j_column]

        row_2 = 6 + i_row
        col_2 = add_to_cell[j_column]

        if value is not None:
            set_cell_value(file2_sheet, row_2, col_2, value)


def compare_cells(file2_sheet, file3_sheet):
    # """Compare a cell from file 2 with a cell from file 3."""
    print("\nNow comparing values between file 2 and file 3:")
    row_2 = int(input("Enter the row number from file 2 to compare: "))
    col_2 = input("Enter the column letter from file 2 to compare: ").upper()

    row_3 = int(input("Enter the row number from file 3 to compare: "))
    col_3 = input("Enter the column letter from file 3 to compare: ").upper()

    value_2 = get_cell_value(file2_sheet, row_2, col_2)
    value_3 = get_cell_value(file3_sheet, row_3, col_3)

    print(f"\nValue from file 2 ({row_2}, {col_2}): {value_2}")
    print(f"Value from file 3 ({row_3}, {col_3}): {value_3}")

    if value_2 == value_3:
        print("The values are identical.")
    else:
        print("The values are different.")

def save_excel(workbook, file_path):
    # """Save the workbook."""
    workbook.save(file_path)
    print(f"Workbook saved to '{file_path}'.")

if __name__ == "__main__":
    # Get file paths for the three files
    file1_path = "X:/AUR/11.2024/01.11.2024/Orders from vendors.xlsx"
    file2_path = "X:/AUR/11.2024/01.11.2024/noul template.xlsx"
    # file3_path = "X:/AUR/11.2024/01.11.2024/Retail Item List.xlsx"

    # Open the three files
    workbook1, sheet1 = open_excel(file1_path, "New vendors and items")
    workbook2, sheet2 = open_excel(file2_path, "Item")
    # workbook3, sheet3 = open_excel(file3_path)

    # if workbook1 and sheet1 and workbook2 and sheet2 and workbook3 and sheet3:
    if workbook1 and sheet1 and workbook2 and sheet2:
        # Copy cells from file 1 to file 2
        copy_cells(sheet1, sheet2)

        # Compare cells between file 2 and file 3
        # compare_cells(sheet2, sheet3)

        # Save changes to file 2
        save_excel(workbook2, file2_path)
