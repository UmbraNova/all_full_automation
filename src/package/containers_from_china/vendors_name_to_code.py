def main(template_file_path="X:/AUR/11.2024/08.11.2024/noul template.xlsx", vendors_file_path="X:/AUR/10.2024/21.10.2024 - Copy/Vendors.xlsx"):
    import openpyxl

    # Load the Excel workbooks
    template_excel = openpyxl.load_workbook(template_file_path)
    vendors_excel = openpyxl.load_workbook(vendors_file_path)

    # Select the active sheets (or specify the sheet name if necessary)
    sheet_1 = template_excel.active
    sheet_2 = vendors_excel.active

    # Create a dictionary to hold names from vendors_excel
    names_dict = {}

    # Loop through column B and G of vendors_excel to populate the dictionary
    for row in range(2, sheet_2.max_row + 1):  # Assuming there's a header
        name_b = sheet_2.cell(row=row, column=2).value  # Column B
        name_g = sheet_2.cell(row=row, column=7).value  # Column G
        value_a = sheet_2.cell(row=row, column=1).value  # Column A

        # Add to dictionary if name in column B or G exists
        if name_b is not None:
            names_dict[name_b] = value_a
        if name_g is not None:
            names_dict[name_g] = value_a

    # Check against names in template_excel and update
    for row in range(2, sheet_1.max_row + 1):  # Assuming there's a header
        name_e = sheet_1.cell(row=row, column=5).value  # Column E
        if name_e in names_dict:
            print(str(name_e))
            sheet_1.cell(row=row, column=5).value = names_dict[name_e]

    # Save the updated template_excel
    # template_excel.save('updated_template_excel.xlsx')
    template_excel.save(template_file_path)

    print("Update complete!")


if __name__ == "__main__":
    main()
