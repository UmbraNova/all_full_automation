# def main(template_file_path="X:/AUR/25.10.2024/updated_template_excel.xlsx"):    
def main(template_file_path="X:/AUR/11.2024/08.11.2024/noul template.xlsx"):    
    import openpyxl
    from openpyxl.styles import PatternFill


    # compara numarul de articol cu barcodul, daca nr de articol nu este regasit in barcode, coloreaza cu rosu

    # Define file paths
    input_file_path = template_file_path  # Change this to your actual input file path
    output_file_path = template_file_path  # Change this to your desired output file path

    # Load the Excel workbook
    workbook = openpyxl.load_workbook(input_file_path)
    sheet = workbook.active  # You can specify the sheet name if needed

    # Define the fill for red background
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    # Get the maximum row number in the sheet
    max_row = sheet.max_row

    # Create a list of values from column AF for quick lookup
    af_values = [str(sheet.cell(row=row, column=32).value) for row in range(1, max_row + 1)]  # Column AF is the 32nd column

    # Iterate through each cell in column O and check against values in column AF
    for row in range(6, max_row + 1):  # Adjust starting row if there's a header
        o_value = str(sheet.cell(row=row, column=15).value)  # Column O is the 15th column
        # Check if o_value is in any of the af_values
        if o_value and not any(o_value in af_value for af_value in af_values):
            # Change the background color to red if the value is not found in AF
            sheet.cell(row=row, column=15).fill = red_fill

    # Save the modified workbook
    print("Done!")
    workbook.save(output_file_path)  # Save the changes to the specified output file


if __name__ == "__main__":
    main()
