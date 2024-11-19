
def main(template_file_path="X:/AUR/containers_auto/noul template.xlsx", containers_folder_path="X:/AUR/containers_auto/containers_folder"):
    import os
    import openpyxl
    import pandas as pd


    # Function to check if any of the keywords are in the worksheet data
    def contains_keywords(sheet, keywords):
        for row in sheet.iter_rows(values_only=True):
            if any(keyword in str(cell).lower() for cell in row if cell is not None for keyword in keywords):
                return True
        return False

    # Function to check if the value in column "B" is a 3-7 digit number
    def is_valid_item_number(value):
        if isinstance(value, int) or (isinstance(value, str) and value.isdigit()):
            return 3 <= len(str(value)) <= 7
        return False

    def invoice_action(sheet, file_path, output_ws, product_row_map, source_columns, dest_columns, dest_column_b8="E"):
        try:
            df = pd.read_excel(file_path, sheet_name='INVOICE', header=4)
        except:
            df = pd.read_excel(file_path, sheet_name='PACKING LIST', header=4)

        b8_value = df.iloc[2, 1]  # Row index 2 (third row), Column index 1 (second column)

        # Selecting only the name of the company
        target_words = ["Address:", "Add:"]
        words = b8_value.split() if b8_value else []
        for target_word in target_words:
            if target_word in words:
                b8_value = ' '.join(words[:words.index(target_word)])[8:-1]
                break

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            item_number = row[1].value  # items column B (index 1)
            if is_valid_item_number(item_number):
                if item_number in product_row_map:
                    output_row = product_row_map[item_number]
                else:
                    output_row = output_ws.max_row + 1
                    product_row_map[item_number] = output_row
                
                # Copy the B8 value to the specified destination column for this product
                output_ws[f"{dest_column_b8}{output_row}"] = b8_value
                
                # output_ws[f"B{output_row}"] = counter

                # Copy data from source columns to destination columns, skipping empty cells
                for src_col, dest_col in zip(source_columns, dest_columns):
                    src_value = sheet[f"{src_col}{row[0].row}"].value
                    if src_value is not None:  # Check if the cell has a value
                        output_ws[f"{dest_col}{output_row}"] = src_value


    def packing_action(sheet, output_ws, product_row_map, source_columns, dest_columns):
        
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            item_number = row[1].value  # items column B (index 1)
            if is_valid_item_number(item_number):
                if item_number in product_row_map:
                    output_row = product_row_map[item_number]
                else:
                    output_row = output_ws.max_row + 1
                    product_row_map[item_number] = output_row
                
                output_ws[f"J{output_row}"] = "MARFA_19"
                output_ws[f"K{output_row}"] = "TVA19B"
                output_ws[f"L{output_row}"] = "CN"
                output_ws[f"R{output_row}"] = "CHINA_200D"
                output_ws[f"S{output_row}"] = "FALSE"
                output_ws[f"V{output_row}"] = "BUC"
                output_ws[f"W{output_row}"] = "BUC"
                output_ws[f"X{output_row}"] = "BUC"
                output_ws[f"Y{output_row}"] = "1"
                output_ws[f"B{output_row}"] = output_row - 5
                # Read values from the source columns, skipping empty cells
                values = []
                for col in source_columns:
                    cell_value = sheet[f"{col}{row[0].row}"].value
                    if cell_value is not None:  # Check if the cell has a value
                        values.append(cell_value)
                    else:
                        values.append(None)  # Keep the structure intact

                # Check if we have enough valid values to perform the calculations
                if len(values) == len(source_columns) and all(isinstance(v, (int, float)) for v in values if v is not None):
                    result_1 = values[0] / values[1]  # qty/bax
                    result_2 = (values[2] / values[0]) * 1000  # brut weight
                    result_3 = (values[3] / values[0]) * 1000  # net weight
                    result_4 = (values[4] / values[0]) * 1000  # cubaj

                    # Store the results in the destination columns
                    results = [result_1, result_2, result_3, result_4]
                    for dest_col, result in zip(dest_columns, results):
                        output_ws[f"{dest_col}{output_row}"] = result

    # Function to process individual files
    def process_file(file_path, output_ws, product_row_map, columns_1, columns_2):
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        file_name = os.path.basename(file_path).lower()  # Get the file name in lowercase
        
        # Check for "invoice" or "barcode" in the file name
        if "invoice" in file_name or "barcode" in file_name:  # Check for substring
            invoice_action(sheet, file_path, output_ws, product_row_map, columns_1[0], columns_1[1])
        elif "list" in file_name:  # Check for "list" in the file name
            packing_action(sheet, output_ws, product_row_map, columns_2[0], columns_2[1])

    def process_folders(base_folder, output_template):
        # Open the template workbook and the active sheet where data will be stored
        output_wb = openpyxl.load_workbook(output_template)
        output_ws = output_wb.active
        
        product_row_map = {}  # Track row for each item number in column "B"
        
        # Source \ Destination
        columns_1 = (["B", "D", "C", "F", "K", "L", "M"], ["O", "C", "D", "AD", "AF", "N", "AI"])  # invoice
        columns_2 = (["E", "G", "H", "I", "J"], ["H", "G", "F", "AC"])  # packing list

        # Traverse through each folder and subfolder
        for root, dirs, files in os.walk(base_folder):
            for file in files:
                if file.endswith('.xlsx'):
                    file_path = os.path.join(root, file)
                    print(f"Processing file: {file}")
                    process_file(file_path, output_ws, product_row_map, columns_1, columns_2)
        
        output_wb.save(output_template)
        print(f"All data processed and saved in {output_template}")

    base_folder = containers_folder_path
    output_template = template_file_path
    process_folders(base_folder, output_template)

if __name__ == "__main__":
    main()
