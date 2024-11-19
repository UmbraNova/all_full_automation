def main(template_file_path="X:/AUR/11.2024/08.11.2024/total_import_noul template.xlsx"):
# def main(template_file_path="X:/AUR/21.10.2024 - Copy/noul template.xlsx"):
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill


    # Provide the full path to your Excel file
    file_path = template_file_path
    output_file_path = template_file_path

    red_fill = PatternFill(start_color="FFC6C2", end_color="FFC6C2", fill_type="solid")
    # Load the Excel workbook and select the active worksheet
    wb = load_workbook(file_path)
    ws = wb.active

    # Loop through each cell in column C (starting from row 1)
    for row in ws.iter_rows(min_row=6, max_col=3, max_row=ws.max_row):
        cell_value = row[2].value  # Column C (index 2)
        
        if cell_value:  # Check if there's text in the cell
            cell_length = len(str(cell_value))
            
            # Debug print to track cell values and lengths
            print(f"Row {row[0].row} - Cell Length: {cell_length} - Value: {cell_value}")
            
            # If the text length is under or equal to 30 characters
            if cell_length <= 30:
                # Place text in the desired location if <30 characters (confirm correct column)
                ws.cell(row=row[0].row, column=33).value = cell_value  # Assuming AG is intended for short text
            else:
                # Split the text without cutting words
                words = cell_value.split()
                part1, part2 = "", ""
    # //////////////////////////////////////////////////////////////////////////

                for i_word in range(len(words)+1):
                    description_text = " ".join(words[:i_word])
                    if len(description_text) <= 30:
                        if description_text[-4:] == "ART." or description_text[-4:] == "ART":
                            part1 = description_text[:-4]
                            part2 = f"ART.{" ".join(words[i_word:])}"
                        else:
                            part1 = description_text
                            part2 = " ".join(words[i_word:])

    # //////////////////////////////////////////////////////////////////////////
                # Write the parts to AG (column 33) and AH (column 34)
                ws.cell(row=row[0].row, column=33).value = part1.strip()  # AG
                ws.cell(row=row[0].row, column=34).value = part2.strip()  # AH
                if len(part2) > 20:
                    ws.cell(row=row[0].row, column=34).fill = red_fill 

    # Save the updated workbook with the full path
    wb.save(output_file_path)

    print(f"File saved successfully to {output_file_path}")


if __name__ == "__main__":
    main()
