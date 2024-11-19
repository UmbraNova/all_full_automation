# compara lista de produse to be imported cu produsele care deja sunt in system
# mereu de re-downloadat fisierul din configuration packages > ITEM > Excel > Export to Excel
# le coloreaza (articolul) cu albastru

import openpyxl
from openpyxl.styles import PatternFill


def main(template_file_path="X:/AUR/11.2024/08.11.2024/2total_import_noul template.xlsx",
         items_in_system_path="X:/AUR/11.2024/08.11.2024/barcodes_in_system.xlsx"):
    
    def compare_and_update(file_one_path, file_two_path, sheet_one_name='Item', sheet_two_name='Barcodes'):
        # Load both workbooks and sheets
        workbook_one = openpyxl.load_workbook(file_one_path)
        workbook_two = openpyxl.load_workbook(file_two_path)
        
        sheet_one = workbook_one[sheet_one_name]
        sheet_two = workbook_two[sheet_two_name]
        
        # Define a blue fill for highlighting matches
        blue_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
        
        # Collect values from file two's specified column into a set for faster lookup
        file_two_values = {sheet_two[f"A{row}"].value for row in range(1, sheet_two.max_row + 1) if sheet_two[f"Z{row}"].value}
        # file_two_values = {sheet_two[f"Z{row}"].value for row in range(1, sheet_two.max_row + 1) if sheet_two[f"Z{row}"].value}
        
        # Iterate through rows in file one to find matches
        for row_one in range(1, sheet_one.max_row + 1):
            cell_one = sheet_one[f"AF{row_one}"]  # Access cell directly for easier modification
            cell_one_value = cell_one.value
            
            # If cell value in file one matches any value from file two, apply blue fill
            if str(cell_one_value) in str(file_two_values):
                cell_one.fill = blue_fill
                print(f"Match found: {cell_one_value} at row {row_one}, marked in blue.")
        
        # Save the updated file_one workbook
        workbook_one.save(file_one_path)
        print("Workbook saved successfully!")

    # Run the comparison and update
    compare_and_update(template_file_path, items_in_system_path)

if __name__ == "__main__":
    main()
